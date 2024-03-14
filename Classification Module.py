import re
import json
import argparse
from openpyxl import load_workbook
import sys

# Function to check the data type of input
def check_input(input_data):
    ip_pattern = r'^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}$'
    domain_pattern = r'^([a-zA-Z0-9]+(-[a-zA-Z0-9]+)*\.)+[a-zA-Z]{2,}$'
    email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'

    if re.match(ip_pattern, input_data):
        return "IP Address"
    elif re.match(domain_pattern, input_data):
        return "Domain Name"
    elif re.match(email_pattern, input_data):
        return "Email Address"
    else:
        return "Unidentified"

# Function to process manually entered data
def process_input(data):
    if data:
        label = check_input(data)
        return {"Input Data": data, "Data Type": label}

# Function to process file input
def process_file(file_path):
    results = []
    if file_path.endswith('.txt'):
        with open(file_path, 'r', encoding='utf-8') as file:
            for line in file:
                line = line.strip()
                label = check_input(line)
                results.append({"Input Data": line, "Data Type": label})
    elif file_path.endswith('.xlsx'):
        wb = load_workbook(file_path)
        sheet = wb.active
        for row in sheet.iter_rows(values_only=True):
            for cell in row:
                label = check_input(str(cell))
                results.append({"Input Data": cell, "Data Type": label})
    return results

# Export results to a specified format
def export_results(results, output_format):
    if output_format == 'json':
        print(json.dumps(results, indent=4))
    elif output_format == 'csv':
        for item in results:
            print(f"{item['Input Data']},{item['Data Type']}")
    elif output_format == 'txt':
        for item in results:
            print(f"Input Data: {item['Input Data']}\nData Type: {item['Data Type']}\n")

# Main function
def main():
    parser = argparse.ArgumentParser(description='Data Classification Tool')
    parser.add_argument('--format', nargs='+', choices=['json', 'csv', 'txt'], default=['txt'], help='Output format(s)')
    parser.add_argument('--input', help='Input file path', type=str)
    parser.add_argument('--stdin', action='store_true', help='Read from stdin')

    args = parser.parse_args()

    if args.stdin:
        results = []
        for line in sys.stdin:
            line = line.strip()
            label = check_input(line)
            results.append({"Input Data": line, "Data Type": label})
    elif args.input:
        results = process_file(args.input)
    else:
        print("Error: Please specify input source (--input or --stdin)")
        return

    for output_format in args.format:
        export_results(results, output_format)

if __name__ == "__main__":
    main()
