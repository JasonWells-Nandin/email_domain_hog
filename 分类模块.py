import re
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
import json

# 检测输入数据类型的函数
def check_input(input_data):
    ip_pattern = r'^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}$'
    domain_pattern = r'^([a-zA-Z0-9]+(-[a-zA-Z0-9]+)*\.)+[a-zA-Z]{2,}$'
    email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'

    if re.match(ip_pattern, input_data):
        return "IP地址"
    elif re.match(domain_pattern, input_data):
        return "域名"
    elif re.match(email_pattern, input_data):
        return "电子邮件地址"
    else:
        return "无法识别"

# 处理手动输入数据的函数
def process_input():
    data = entry.get()  # 获取输入框中的数据
    if data:
        label = check_input(data)  # 检测数据类型
        result_label.config(text="输入数据: {}\n数据类型: {}".format(data, label))  # 显示检测结果

# 处理选择文件的函数
def browse_file():
    file_path = filedialog.askopenfilename()  # 弹出文件选择对话框，返回文件路径
    if file_path:
        if file_path.endswith('.txt'):  # 如果是文本文件
            with open(file_path, 'r', encoding='utf-8') as file:  # 以 utf-8 编码打开文件
                output = ""
                results = []
                for line in file:
                    line = line.strip()  # 去除行末尾的换行符和空白符
                    label = check_input(line)  # 检测当前行数据的类型
                    results.append({"输入数据": line, "数据类型": label})
                    output += "输入数据: {}\n数据类型: {}\n\n".format(line, label)
                result_label.config(text=output)  # 显示检测结果
                return results
        elif file_path.endswith('.xlsx'):  # 如果是 Excel 文件
            wb = load_workbook(file_path)
            sheet = wb.active
            output = ""
            results = []
            for row in sheet.iter_rows(values_only=True):
                for cell in row:
                    label = check_input(str(cell))  # 检测当前单元格数据的类型
                    results.append({"输入数据": cell, "数据类型": label})
                    output += "输入数据: {}\n数据类型: {}\n\n".format(cell, label)
            result_label.config(text=output)  # 显示检测结果
            return results
        else:
            result_label.config(text="不支持的文件类型")

# 导出结果到指定路径的 JSON 文件
def export_to_json(results):
    file_path = filedialog.asksaveasfilename(defaultextension=".json", filetypes=[("JSON Files", "*.json")])
    if file_path:
        with open(file_path, "w", encoding="utf-8") as json_file:
            json.dump(results, json_file, ensure_ascii=False, indent=4)

# 主函数
def main():
    root = tk.Tk()  # 创建主窗口
    root.title("数据分类工具")  # 设置窗口标题

    frame = tk.Frame(root)  # 创建框架
    frame.pack(padx=10, pady=10)  # 设置框架的填充

    entry_label = tk.Label(frame, text="输入要检测的数据:")  # 创建标签
    entry_label.grid(row=0, column=0)  # 放置标签

    global entry  # 声明entry为全局变量
    entry = tk.Entry(frame, width=50)  # 创建输入框
    entry.grid(row=0, column=1)  # 放置输入框

    process_button = tk.Button(frame, text="检测", command=process_input)  # 创建按钮
    process_button.grid(row=0, column=2, padx=5)  # 放置按钮

    file_button = tk.Button(frame, text="选择文件", command=lambda: handle_file_selection(browse_file()))  # 创建按钮，并绑定选择文件和导出 JSON 的函数
    file_button.grid(row=0, column=3, padx=5)  # 放置按钮

    global result_label  # 声明result_label为全局变量
    result_label = tk.Label(root, text="", justify="left")  # 创建标签
    result_label.pack(pady=10)  # 放置标签

    root.mainloop()  # 运行主事件循环

# 处理CLI模式下的输入数据的函数
def cli_mode():
    while True:
        data = input("请输入要检测的数据 (输入 q 退出): ")
        if data.lower() == 'q':
            break
        label = check_input(data)
        print("输入数据:", data)
        print("数据类型:", label)
        print()

# 处理文件选择的结果
def handle_file_selection(results):
    if results:
        export_choice = tk.messagebox.askyesno("导出JSON", "要导出为JSON文件吗?")
        if export_choice:
            export_to_json(results)

if __name__ == "__main__":
    main()  # 调用主函数启动GUI程序
    print("\n进入CLI模式：\n")
    cli_mode()  # 进入CLI模式
